#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import io
import itertools
import json
import logging
import re
import sys
import zipfile
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator, Optional
from urllib.parse import urlparse

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required. Install it with:\n    pip install openpyxl")

csv.field_size_limit(10 * 1024 * 1024)

log = logging.getLogger("intelxtract")

# Adjust here if your export/combolist format differs
PROFESSIONAL_SHEET = "Professional Leaks"
PERSONAL_SHEET = "Personal Leaks"
SUMMARY_SHEET = "Summary"

FIELD_DELIMITERS = [":", ";", "|", "\t", ","]
TABLE_DELIMITERS = [",", "\t", ";", "|"]  # ':' is handled per-line (email:pass, url:user:pass)
EMAIL_RE = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
USERNAME_RE = re.compile(r"^[A-Za-z0-9][A-Za-z0-9._\-]{1,63}$")
HOST_RE = re.compile(r"^[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")
SCHEME_RE = re.compile(r"^[a-zA-Z][a-zA-Z0-9+.\-]*://")

# Hashed password shapes
HEX_ONLY_RE = re.compile(r"^[0-9a-fA-F]+$")
HASH_HEX_LENGTHS = {32, 40, 56, 64, 96, 128}  # MD5/NTLM, SHA1, SHA224, SHA256, SHA384, SHA512
CRYPT_RE = re.compile(r"^\$(2[abxy]?|1|5|6|y|argon2[id]?|scrypt|pbkdf2)\$")  # bcrypt/sha-crypt/argon2

# Values that are NOT credentials (record GUIDs/ObjectIds, dates, amounts, phone numbers, IP addresses)
UUID_RE = re.compile(r"^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$")
AMOUNT_RE = re.compile(r"\d{1,9}[.,]\d{1,2}$")
ISO_DT_RE = re.compile(r"^\d{4}-\d{2}-\d{2}([T ]\d{2}:\d{2}(:\d{2})?(\.\d+)?(Z|[+\-]\d{2}:?\d{2})?)?$")
SLASH_DATE_RE = re.compile(r"^\d{1,2}[/.]\d{1,2}[/.]\d{2,4}$")
PHONE_ALLOWED_RE = re.compile(r"^[\d+().\-\s]+$")  # only phone-ish chars
PHONE_SEP_RE = re.compile(r"[-+.()\s]")            # must carry phone formatting
IPV4_RE = re.compile(r"^(\d{1,3})\.(\d{1,3})\.(\d{1,3})\.(\d{1,3})$")
IPV6ISH_RE = re.compile(r"^[0-9a-fA-F:]+$")

MAX_PASSWORD_LEN = 128
TEXT_EXTENSIONS = {".txt", ".csv", ".log", ".tsv", ".dat", ".sql", ".json", ""}
INDEX_NAME_HINTS = ("index", "info", "report", "manifest", "metadata", "results")
SAMPLE_ROWS = 200

DATE_FORMATS = (
    "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d",
    "%Y%m%d", "%d/%m/%Y", "%m/%d/%Y",
)
DATE_IN_NAME_RE = re.compile(r"(20\d{2})[-_.]?(\d{2})[-_.]?(\d{2})")

EXCEL_HEADERS = ["Email / Username", "Password", "Leak Name", "Leak Date", "Source URL/Host"]
ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")

MAX_EXCEL_DATA_ROWS = 1_000_000
PREVIEW_ROWS = 1_000

# Data model
@dataclass
class Credential:
    identity: str
    password: str
    leak_name: str
    leak_date: Optional[datetime]
    source_file: str
    is_email: bool
    host: Optional[str] = None
    category: str = "unrelated"

    @property
    def dedup_key(self) -> tuple[str, str]:
        return (self.identity.lower(), self.password)

# Value classifiers
def _is_hash(v: str) -> bool:
    """A hashed password (still a leaked credential -> keep)."""
    if CRYPT_RE.match(v):
        return True
    if HEX_ONLY_RE.match(v) and len(v) in HASH_HEX_LENGTHS:
        return any(c in "abcdefABCDEF" for c in v)  # real hashes carry hex letters, not all-digits
    return False


def _looks_like_date(v: str) -> bool:
    return bool(ISO_DT_RE.match(v) or SLASH_DATE_RE.match(v))


def _looks_like_phone(v: str) -> bool:
    """Formatted phone numbers (with +, -, ., (), spaces). Bare digit runs are
    NOT treated as phones, so numeric passwords like '12345678' survive."""
    if not PHONE_ALLOWED_RE.match(v) or not PHONE_SEP_RE.search(v):
        return False
    return 7 <= len(re.sub(r"\D", "", v)) <= 15


def _looks_like_ip(v: str) -> bool:
    m = IPV4_RE.match(v)
    if m and all(0 <= int(g) <= 255 for g in m.groups()):
        return True
    return bool(IPV6ISH_RE.match(v)) and v.count(":") >= 2


def _is_garbage(v: str) -> bool:
    """Values that are NOT credentials: record GUIDs/ObjectIds, long non-hash
    hex ids, dates, amounts, phone numbers, IPs."""
    if _is_hash(v):
        return False
    if UUID_RE.match(v):                          # IntelX record GUID
        return True
    if HEX_ONLY_RE.match(v) and len(v) >= 16:     # 24=ObjectId, or other long hex ids
        return True
    if _looks_like_date(v):
        return True
    if AMOUNT_RE.fullmatch(v):
        return True
    if _looks_like_phone(v):
        return True
    if _looks_like_ip(v):
        return True
    return False

# Low-level helpers
def _decode(raw: bytes) -> str:
    for enc in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _xlsx_safe(value):
    return ILLEGAL_XLSX_RE.sub("", value) if isinstance(value, str) else value


def parse_date(value) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    value = str(value).strip()
    if not value:
        return None
    if re.fullmatch(r"\d{10}", value):
        try:
            return datetime.fromtimestamp(int(value), tz=timezone.utc).replace(tzinfo=None)
        except (ValueError, OverflowError, OSError):
            pass
    if re.fullmatch(r"\d{13}", value):
        try:
            return datetime.fromtimestamp(int(value) / 1000, tz=timezone.utc).replace(tzinfo=None)
        except (ValueError, OverflowError, OSError):
            pass
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        pass
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    m = DATE_IN_NAME_RE.search(value)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            return None
    return None


def _clean_password(pwd: str, allow_spaces: bool = False, strict: bool = False) -> Optional[str]:
    pwd = pwd.strip()
    if not pwd or len(pwd) > MAX_PASSWORD_LEN:
        return None
    if not allow_spaces and any(c.isspace() for c in pwd):
        return None
    if not any(c.isalnum() for c in pwd):
        return None
    if strict:
        if "," in pwd:          # a field still containing a comma is a CSV record, not a password
            return None
        if _is_garbage(pwd):    # GUIDs/ObjectIds/dates/amounts (hashes pass)
            return None
    return pwd


def extract_host(url: Optional[str]) -> Optional[str]:
    if not url:
        return None
    url = url.strip()
    try:
        if SCHEME_RE.match(url):
            host = urlparse(url).hostname
        else:
            token = re.split(r"[/:\s\\]", url, 1)[0]
            host = token if HOST_RE.match(token) else None
    except ValueError:
        host = None
    return host.lower() if host else None

# Credential extraction
def parse_line(line: str) -> Optional[tuple[Optional[str], str, str, bool]]:
    line = line.strip().strip("\ufeff")
    if not line or len(line) > 4096:
        return None

    m = EMAIL_RE.search(line)
    if m:
        email = m.group(0)
        url = line[:m.start()].rstrip()
        url = url.rstrip("".join(FIELD_DELIMITERS)).strip() or None
        after = line[m.end():].lstrip()
        if after[:1] in FIELD_DELIMITERS:
            sep = after[0]
            field = after[1:].split(sep, 1)[0]
        else:
            field = after
        pwd = _clean_password(field.strip(), strict=True)
        if pwd is None:
            return None
        return (url, email.lower(), pwd, True)

    if "://" in line:
        parts = line.split(":")
        if len(parts) >= 3:
            user, pwd = parts[-2].strip(), parts[-1].strip()
            url = ":".join(parts[:-2]).strip() or None
            pwd_c = _clean_password(pwd, strict=True)
            if USERNAME_RE.match(user) and pwd_c is not None:
                return (url, user, pwd_c, False)
        return None

    for d in FIELD_DELIMITERS:
        if d in line:
            parts = line.split(d)
            if len(parts) == 2:
                user, pwd = parts[0].strip(), parts[1].strip()
                pwd_c = _clean_password(pwd, strict=True)
                if USERNAME_RE.match(user) and pwd_c is not None:
                    return (None, user, pwd_c, False)
            return None
    return None


def _pwd_score(values: list[str]) -> float:
    vals = [v.strip() for v in values if v.strip()]
    if not vals:
        return 0.0
    good = 0
    for v in vals:
        if not (4 <= len(v) <= MAX_PASSWORD_LEN) or " " in v:
            continue
        if _is_garbage(v):
            continue
        if _is_hash(v) or any(ch.isalpha() for ch in v):
            good += 1
    score = good / len(vals)
    if len(set(vals)) / len(vals) < 0.3:   # low cardinality => categorical
        score *= 0.3
    return score


def _column_is_garbage(values: list[str]) -> bool:
    vals = [v.strip() for v in values if v.strip()]
    if not vals:
        return True
    bad = sum(1 for v in vals if _is_garbage(v))
    return bad / len(vals) >= 0.7


def _iter_tabular(text: str, delim: str, width: int):
    sample = list(itertools.islice(csv.reader(io.StringIO(text), delimiter=delim), SAMPLE_ROWS))
    sample = [r for r in sample if len(r) >= 2]
    if not sample:
        return

    def col(i):
        return [r[i] for r in sample if len(r) > i]

    def frac(i, pred):
        vals = col(i)
        return (sum(1 for v in vals if pred(v.strip())) / len(vals)) if vals else 0.0

    email_col = max(range(width), key=lambda i: frac(i, lambda v: bool(EMAIL_RE.fullmatch(v))), default=None)
    if email_col is None or frac(email_col, lambda v: bool(EMAIL_RE.fullmatch(v))) < 0.3:
        log.debug("Skipping table with no email column (%d cols, delim %r).", width, delim)
        return
    id_col = email_col

    candidates = [i for i in range(width) if i != id_col]
    pwd_col = None
    if width == 2 and candidates:
        i = candidates[0]
        if not _column_is_garbage(col(i)):
            pwd_col = i
    else:
        scored = [(i, _pwd_score(col(i))) for i in candidates]
        if scored:
            bi, bs = max(scored, key=lambda t: t[1])
            if bs >= 0.5:
                pwd_col = bi
    if pwd_col is None:
        log.debug("Skipping file with no password column (%d cols, delim %r).", width, delim)
        return

    url_col = None
    cands = [(i, frac(i, lambda v: ("://" in v) or bool(HOST_RE.match(v))))
             for i in range(width) if i not in (id_col, pwd_col)]
    if cands:
        i, f = max(cands, key=lambda t: t[1])
        if f >= 0.5:
            url_col = i

    need = max(id_col, pwd_col, url_col or 0)
    for row in csv.reader(io.StringIO(text), delimiter=delim):
        if len(row) <= need:
            continue
        identity = row[id_col].strip()
        if not EMAIL_RE.fullmatch(identity):
            continue
        identity = identity.lower()
        pwd = _clean_password(row[pwd_col].strip(), allow_spaces=True, strict=True)
        if pwd is None:
            continue
        url = row[url_col].strip() if url_col is not None else None
        yield (url, identity, pwd, True)


def _iter_header_csv(text: str):
    reader = csv.reader(io.StringIO(text))
    try:
        header = next(reader)
    except StopIteration:
        return
    cols = [h.strip().lower() for h in header]

    def find(*names):
        for i, c in enumerate(cols):
            if c in names:
                return i
        for i, c in enumerate(cols):
            if any(n in c for n in names):
                return i
        return None

    email_i = find("email", "e-mail", "mail", "login", "username", "user", "usuario")
    pass_i = find("password", "pass", "pwd", "senha", "hash")
    url_i = find("url", "host", "domain", "website", "site")
    if email_i is None or pass_i is None:
        return
    for row in reader:
        if len(row) <= max(email_i, pass_i):
            continue
        identity = row[email_i].strip()
        pwd = _clean_password(row[pass_i].strip(), allow_spaces=True)  # header trusted
        if not identity or pwd is None:
            continue
        url = row[url_i].strip() if (url_i is not None and len(row) > url_i) else None
        if EMAIL_RE.fullmatch(identity):
            yield (url, identity.lower(), pwd, True)
        elif USERNAME_RE.match(identity):
            yield (url, identity, pwd, False)


def _table_shape(sample_lines: list[str]) -> Optional[tuple[str, int]]:
    best = None
    for d in TABLE_DELIMITERS:
        counts = [ln.count(d) for ln in sample_lines]
        present = [c for c in counts if c >= 1]
        if len(present) < max(3, int(len(sample_lines) * 0.7)):
            continue
        modal, freq = Counter(c + 1 for c in present).most_common(1)[0]
        if modal >= 2 and freq >= len(present) * 0.7:
            if best is None or freq > best[2]:
                best = (d, modal, freq)
    return (best[0], best[1]) if best else None


def iter_credentials_from_text(text: str):
    lines = [ln for ln in text.splitlines() if ln.strip()]
    if not lines:
        return
    sample = lines[:SAMPLE_ROWS]
    first = sample[0].lower()
    header_like = (
        any(k in first for k in ("password", "pass", "pwd", "senha", "hash"))
        and any(k in first for k in ("email", "e-mail", "mail", "login", "user", "usuario"))
        and any(sep in sample[0] for sep in (",", ";", "\t"))
    )
    if header_like:
        yield from _iter_header_csv(text)
        return

    shape = _table_shape(sample)
    if shape:
        yield from _iter_tabular(text, shape[0], shape[1])
        return

    for ln in lines:
        try:
            res = parse_line(ln)
        except Exception:
            res = None
        if res:
            yield res

# Metadata (leak name + date) resolution
def _key_variants(value: str) -> set[str]:
    value = value.strip().lower()
    p = Path(value)
    return {v for v in {value, p.name, p.stem} if v}


def _merge_csv_index(text: str, index: dict) -> bool:
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(io.StringIO(text), dialect=dialect)
    if not reader.fieldnames:
        return False
    cols = {f.strip().lower(): f for f in reader.fieldnames}

    def pick(*names):
        for n in names:
            if n in cols:
                return cols[n]
        for low, orig in cols.items():
            if any(n in low for n in names):
                return orig
        return None

    key_col = pick("systemid", "system id", "id", "file", "filename", "name")
    name_col = pick("name", "title", "filename", "file")
    date_col = pick("date", "added", "indexed", "datetime", "time")
    if key_col is None:
        return False

    matched = False
    for row in reader:
        key = (row.get(key_col) or "").strip()
        if not key:
            continue
        entry = {
            "name": (row.get(name_col) or "").strip() if name_col else "",
            "date": parse_date(row.get(date_col)) if date_col else None,
        }
        for k in _key_variants(key):
            index.setdefault(k, entry)
        matched = True
    return matched


def _merge_json_index(text: str, index: dict) -> bool:
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        return False
    records: list = []
    if isinstance(data, list):
        records = data
    elif isinstance(data, dict):
        for v in data.values():
            if isinstance(v, list):
                records = v
                break
        if not records and all(not isinstance(v, (list, dict)) for v in data.values()):
            records = [data]

    matched = False
    for rec in records:
        if not isinstance(rec, dict):
            continue
        low = {str(k).lower(): v for k, v in rec.items()}
        key = (low.get("systemid") or low.get("system id") or low.get("id")
               or low.get("file") or low.get("filename") or low.get("name"))
        if not key:
            continue
        name = low.get("name") or low.get("title") or low.get("file") or low.get("filename") or ""
        date = parse_date(low.get("date") or low.get("added") or low.get("indexed")
                          or low.get("datetime") or low.get("time"))
        entry = {"name": str(name).strip(), "date": date}
        for kv in _key_variants(str(key)):
            index.setdefault(kv, entry)
        matched = True
    return matched


def load_metadata_index(zf: zipfile.ZipFile) -> tuple[dict, set]:
    index: dict = {}
    consumed: set = set()
    for info in zf.infolist():
        if info.is_dir():
            continue
        base = Path(info.filename).name.lower()
        ext = Path(info.filename).suffix.lower()
        if not any(h in base for h in INDEX_NAME_HINTS) or ext not in (".csv", ".tsv", ".json"):
            continue
        try:
            text = _decode(zf.read(info.filename))
        except (KeyError, zipfile.BadZipFile):
            continue
        matched = (_merge_json_index(text, index) if ext == ".json" or text.lstrip().startswith(("{", "["))
                   else _merge_csv_index(text, index))
        if matched:
            consumed.add(info.filename)
    if index:
        log.info("Loaded metadata for %d item(s) from index file(s).", len(index))
    return index, consumed


def _clean_leak_name(filename: str) -> str:
    stem = Path(filename).stem
    stem = DATE_IN_NAME_RE.sub("", stem)
    stem = re.sub(r"[_\-]+", " ", stem).strip(" -_")
    stem = re.sub(r"\s{2,}", " ", stem)
    return stem or Path(filename).stem


def _zip_datetime(info: zipfile.ZipInfo) -> Optional[datetime]:
    try:
        dt = datetime(*info.date_time)
    except (ValueError, TypeError):
        return None
    return None if dt.year <= 1980 else dt


def resolve_metadata(info: zipfile.ZipInfo, meta_index: dict) -> tuple[str, Optional[datetime]]:
    base = Path(info.filename).name
    entry = meta_index.get(base.lower()) or meta_index.get(Path(base).stem.lower())
    name = (entry.get("name") if entry else "") or ""
    date = entry.get("date") if entry else None
    if not name:
        name = _clean_leak_name(base)
    if date is None:
        date = parse_date(base)
    if date is None:
        date = _zip_datetime(info)
    return name, date

# Domain classification
def _domain_matches(email: str, domain: str) -> bool:
    try:
        edomain = email.rsplit("@", 1)[1].lower()
    except IndexError:
        return False
    return edomain == domain or edomain.endswith("." + domain)


def _host_matches(host: str, domain: str) -> bool:
    host = host.lower()
    return host == domain or host.endswith("." + domain)


def compute_relevant_files(zf, meta_index, consumed, domain, domain_label) -> set:
    relevant: set = set()
    if not domain:
        return relevant
    needles = {domain}
    if domain_label and len(domain_label) >= 4:
        needles.add(domain_label)
    for info in zf.infolist():
        if info.is_dir() or info.filename in consumed:
            continue
        if Path(info.filename).suffix.lower() not in TEXT_EXTENSIONS:
            continue
        base = Path(info.filename).name
        leak_name, _ = resolve_metadata(info, meta_index)
        haystack = f"{base} {leak_name}".lower()
        if any(n in haystack for n in needles):
            relevant.add(base.lower())
    return relevant


def classify(c: Credential, domain: Optional[str], relevant_files: set) -> str:
    if domain and c.is_email and _domain_matches(c.identity, domain):
        return "professional"
    if domain and c.host and _host_matches(c.host, domain):
        return "appuser"
    if c.source_file.lower() in relevant_files:
        return "appuser"
    return "unrelated"

# Streaming collection
def iter_zip_credentials(zf, meta_index, consumed) -> Iterator[Credential]:
    for info in zf.infolist():
        if info.is_dir() or info.filename in consumed:
            continue
        if Path(info.filename).suffix.lower() not in TEXT_EXTENSIONS:
            continue
        try:
            text = _decode(zf.read(info.filename))
        except (KeyError, zipfile.BadZipFile) as e:
            log.warning("Could not read %s: %s", info.filename, e)
            continue
        if not text.strip():
            continue
        leak_name, leak_date = resolve_metadata(info, meta_index)
        base = Path(info.filename).name
        for url, identity, password, is_email in iter_credentials_from_text(text):
            yield Credential(
                identity=identity, password=password, leak_name=leak_name,
                leak_date=leak_date, source_file=base, is_email=is_email,
                host=extract_host(url),
            )


def _keep_over(candidate: Credential, current: Credential) -> bool:
    """When the same (identity, password) appears in several leaks, keep the
    OLDEST occurrence. Records without a date are treated as newest, so a dated
    record is preferred over an undated one."""
    return (candidate.leak_date or datetime.max) < (current.leak_date or datetime.max)


def sort_recent_first(creds: list[Credential]) -> list[Credential]:
    creds = sorted(creds, key=lambda c: c.identity.lower())
    creds = sorted(creds, key=lambda c: c.leak_date or datetime.min, reverse=True)
    return creds


def collect(zip_path: Path, domain: Optional[str], include_unrelated: bool, out_stem: Path):
    stats = Counter()
    related: dict[tuple[str, str], Credential] = {}
    unrelated_csv = None
    unrelated_seen: set[tuple[str, str]] = set()
    writer = fh = None

    with zipfile.ZipFile(zip_path) as zf:
        meta_index, consumed = load_metadata_index(zf)

        if domain is None:
            counter: Counter = Counter()
            for c in iter_zip_credentials(zf, meta_index, consumed):
                if c.is_email:
                    counter[c.identity.rsplit("@", 1)[1]] += 1
            domain = counter.most_common(1)[0][0] if counter else None
            if domain:
                log.warning("No --domain given; inferred '%s' (unreliable on multi-breach "
                            "exports). Pass --domain to be sure.", domain)

        domain_label = domain.split(".")[0] if domain else None
        relevant_files = compute_relevant_files(zf, meta_index, consumed, domain, domain_label)
        if relevant_files:
            log.info("%d source file(s) look domain-specific by name.", len(relevant_files))

        if include_unrelated:
            unrelated_csv = out_stem.with_name(out_stem.name + "_unrelated.csv")
            fh = open(unrelated_csv, "w", newline="", encoding="utf-8")
            writer = csv.writer(fh)
            writer.writerow(EXCEL_HEADERS)

        for c in iter_zip_credentials(zf, meta_index, consumed):
            cat = classify(c, domain, relevant_files)
            stats[cat] += 1
            if cat in ("professional", "appuser"):
                c.category = cat
                cur = related.get(c.dedup_key)
                if cur is None or _keep_over(c, cur):
                    related[c.dedup_key] = c
            elif include_unrelated:
                if c.dedup_key not in unrelated_seen:
                    unrelated_seen.add(c.dedup_key)
                    writer.writerow([
                        c.identity, c.password, c.leak_name,
                        c.leak_date.strftime("%Y-%m-%d") if c.leak_date else "",
                        c.host or "",
                    ])

    if fh:
        fh.close()

    professional = sort_recent_first([c for c in related.values() if c.category == "professional"])
    personal = sort_recent_first([c for c in related.values() if c.category == "appuser"])
    return professional, personal, domain, stats, unrelated_csv

# Output
def _row_values(c: Credential) -> list:
    return [_xlsx_safe(c.identity), _xlsx_safe(c.password), _xlsx_safe(c.leak_name),
            c.leak_date, _xlsx_safe(c.host or "")]


def _autosize(ws, max_width: int = 70) -> None:
    widths: dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            s = cell.value.strftime("%Y-%m-%d") if isinstance(cell.value, datetime) else str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), len(s))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max_width, w + 2)


def _style_header_row(ws, row_idx: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(EXCEL_HEADERS) + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.fill, cell.font = fill, font
        cell.alignment = Alignment(vertical="center")


def _slug(name: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")


def write_bucket(wb, sheet_name: str, creds: list[Credential], out_stem: Path) -> Optional[Path]:
    ws = wb.create_sheet(sheet_name)
    date_col = EXCEL_HEADERS.index("Leak Date") + 1

    if len(creds) <= MAX_EXCEL_DATA_ROWS:
        ws.append(EXCEL_HEADERS)
        _style_header_row(ws, 1)
        for c in creds:
            ws.append(_row_values(c))
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=date_col).number_format = "yyyy-mm-dd"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        _autosize(ws)
        return None

    csv_path = out_stem.with_name(f"{out_stem.name}_{_slug(sheet_name)}.csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(EXCEL_HEADERS)
        for c in creds:
            w.writerow([c.identity, c.password, c.leak_name,
                        c.leak_date.strftime("%Y-%m-%d") if c.leak_date else "",
                        c.host or ""])
    ws["A1"] = (f"{len(creds):,} rows exceed Excel's per-sheet limit. "
                f"Full data written to: {csv_path.name}")
    ws["A1"].font = Font(bold=True, color="B00020")
    ws.append([])
    ws.append(EXCEL_HEADERS)
    _style_header_row(ws, ws.max_row)
    for c in creds[:PREVIEW_ROWS]:
        ws.append(_row_values(c))
    log.warning("%s has %d rows; written to CSV %s with a %d-row preview in Excel.",
                sheet_name, len(creds), csv_path.name, PREVIEW_ROWS)
    return csv_path


def add_summary_sheet(wb, professional, personal, domain, stats, args, spill_paths, unrelated_csv) -> None:
    ws = wb.create_sheet(SUMMARY_SHEET)
    ws["A1"] = "IntelXtract - Summary"
    ws["A1"].font = Font(bold=True, size=14)

    included = []
    if not args.only_personal_emails:
        included += professional
    if not args.only_domain_emails:
        included += personal

    rows = [
        ("Searched domain", domain or "(unknown / not provided)"),
        ("Professional (domain) leaks", len(professional)),
        ("Personal (app users on domain)", len(personal)),
        ("Unrelated lines in export (excluded)", stats.get("unrelated", 0)),
    ]
    dates = [c.leak_date for c in included if c.leak_date]
    if dates:
        rows.append(("Earliest leak date", min(dates).strftime("%Y-%m-%d")))
        rows.append(("Most recent leak date", max(dates).strftime("%Y-%m-%d")))
    if unrelated_csv:
        rows.append(("Unrelated CSV", unrelated_csv.name))
    for p in spill_paths:
        rows.append(("Spilled to CSV (too big for Excel)", p.name))

    r = 3
    for key, val in rows:
        ws.cell(row=r, column=1, value=key).font = Font(bold=True)
        ws.cell(row=r, column=2, value=val)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Top leaks by credential count").font = Font(bold=True)
    r += 1
    for name, count in Counter(c.leak_name for c in included).most_common(15):
        ws.cell(row=r, column=1, value=_xlsx_safe(name))
        ws.cell(row=r, column=2, value=count)
        r += 1

    r += 1
    ws.cell(row=r, column=1,
            value=("Passwords (plaintext) and password hashes (MD5/SHA/bcrypt) are both kept "
                   "as credentials. Non-credential values -- record GUIDs, ObjectIds, dates "
                   "and amounts -- are excluded. 'Unrelated' lines are credentials in the same "
                   "exported files but not tied to the domain (see --include-unrelated)."))
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 30


def build_workbook(professional, personal, domain, stats, args, out_stem, unrelated_csv) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    spill_paths = []
    if not args.only_personal_emails:
        sp = write_bucket(wb, PROFESSIONAL_SHEET, professional, out_stem)
        if sp:
            spill_paths.append(sp)
    if not args.only_domain_emails:
        sp = write_bucket(wb, PERSONAL_SHEET, personal, out_stem)
        if sp:
            spill_paths.append(sp)
    add_summary_sheet(wb, professional, personal, domain, stats, args, spill_paths, unrelated_csv)
    return wb

# CLI
def parse_args(argv=None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        prog="intelxtract.py",
        description="IntelXtract: parse an IntelX domain-search export .zip into an Excel "
                    "report of leaked credentials tied to the domain (professional vs app users).",
    )
    p.add_argument("zip", metavar="EXPORT_ZIP", help="Path to the IntelX export .zip")
    p.add_argument("-o", "--output", help="Output .xlsx path (default: <zip name>_leaks.xlsx)")
    p.add_argument("-d", "--domain",
                   help="Domain searched on IntelX (e.g. example.com). Strongly recommended; "
                        "if omitted it is inferred from the most common email domain.")
    p.add_argument("--include-unrelated", action="store_true",
                   help="Also export credentials NOT tied to the domain (co-resident in the "
                        "exported dumps) to a separate CSV. Can be very large.")
    group = p.add_mutually_exclusive_group()
    group.add_argument("--only-domain-emails", action="store_true",
                       help="Export only professional leaks (emails on the searched domain).")
    group.add_argument("--only-personal-emails", action="store_true",
                       help="Export only personal leaks (app users on the domain).")
    p.add_argument("-v", "--verbose", action="store_true", help="Verbose logging.")
    return p.parse_args(argv)


def main(argv=None) -> int:
    args = parse_args(argv)
    logging.basicConfig(level=logging.DEBUG if args.verbose else logging.INFO,
                        format="%(levelname)s: %(message)s")

    zip_path = Path(args.zip)
    if not zip_path.is_file():
        log.error("File not found: %s", zip_path)
        return 2
    if not zipfile.is_zipfile(zip_path):
        log.error("Not a valid .zip file: %s", zip_path)
        return 2

    out_path = Path(args.output) if args.output else zip_path.with_name(zip_path.stem + "_leaks.xlsx")
    out_stem = out_path.with_suffix("")
    domain = args.domain.lower().lstrip("@") if args.domain else None

    professional, personal, domain, stats, unrelated_csv = collect(
        zip_path, domain, args.include_unrelated, out_stem)

    log.info("Domain: %s | professional=%d  app-users=%d  unrelated(excluded)=%d",
             domain, len(professional), len(personal), stats.get("unrelated", 0))

    wb = build_workbook(professional, personal, domain, stats, args, out_stem, unrelated_csv)
    wb.save(out_path)

    print(f"\nDone. Wrote: {out_path}")
    if not args.only_personal_emails:
        print(f"  Professional (domain) leaks:      {len(professional):,}")
    if not args.only_domain_emails:
        print(f"  Personal (app users on domain):   {len(personal):,}")
    print(f"  Unrelated lines in export:        {stats.get('unrelated', 0):,} "
          f"({'written to ' + unrelated_csv.name if unrelated_csv else 'excluded -- use --include-unrelated'})")
    print("  NOTE: this file contains plaintext credentials -- handle it as sensitive data.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
